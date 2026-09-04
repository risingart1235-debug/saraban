"""printfmt.py — จัดหน้าทะเบียนรับ/ทะเบียนส่ง ให้พร้อมพิมพ์ A4 แนวนอน

ปัญหาที่แก้
-----------
ทะเบียนบน Google Sheets ตั้งต้นมาเป็น Arial ๑๐ คอลัมน์กว้างเท่ากันหมด ๑๐๐ พิกเซล
พอสั่งพิมพ์จะได้กระดาษที่ช่อง "เรื่อง" ถูกตัดหาย ตัวหนังสือเล็กจนอ่านไม่ออก
และหัวตารางโผล่แค่หน้าแรก — ต้องมานั่งปรับมือใหม่ทุกครั้งที่จะพิมพ์

ไฟล์นี้ทำให้สองอย่าง

๑. ไฟล์ PDF พร้อมพิมพ์ (นี่คือของที่เอาไปพิมพ์จริง)
   วาดหน้ากระดาษเอง จึงคุมได้ทุกอย่างที่กูเกิลคุมไม่ได้:
     * TH Sarabun New ของจริง ขนาด ๑๖ · A4 แนวนอน · หัวตารางซ้ำทุกหน้า · เลขหน้า
     * ขึ้นบรรทัดใหม่ที่ "ขอบคำ" ไม่ตัดกลางคำ — กูเกิลทำข้อนี้ไม่ได้เลย
       (ภาษาไทยเขียนติดกันไม่มีเว้นวรรค กูเกิลจึงตัดตรงไหนก็ได้ ออกมาเป็น
        "เลขทะเบี/ยนรับ" "กลุ่มส่งเสริมการจัดการศึ/กษา" และสั่งแก้ไม่ได้)

๒. หน้าตาบนชีตเองให้อ่านง่ายและพร้อมพิมพ์จากกูเกิลด้วย (เผื่อสั่งพิมพ์จากมือถือ)
     * ตัวอักษรสารบรรณขนาด ๑๖ หัวตารางตัวหนา พื้นเทาอ่อน ตรึงไว้แถวบน
     * ความกว้างคอลัมน์พอดีหน้า A4 แนวนอน ไม่มีคอลัมน์ไหนตกขอบ
     * เลขในทะเบียนเป็นเลขไทยทั้งแผ่น ไม่ปนอารบิกอย่างที่เป็นอยู่
     * ให้ลิงก์สั่งพิมพ์ที่ตั้งค่าหน้ากระดาษมาให้แล้ว (API ตั้งในชีตไม่ได้
       ต้องฝากไปกับลิงก์) เก็บเป็นบุ๊กมาร์กไว้กดดูได้ทันที

ที่วัดได้กับทะเบียนจริง (๔๐๕ เรื่อง): ไฟล์ PDF ๕๕ หน้า / พิมพ์ผ่านกูเกิล ๕๔ หน้า
ต่างกันหน้าเดียว — ตัดคำให้ถูกต้องแทบไม่เปลืองกระดาษเพิ่มเลย

ไฟล์ .xlsx ในเครื่อง (โหมด local) จัดหน้าให้ด้วย และฝังค่าหน้ากระดาษไว้ในไฟล์ได้จริง
โปรแกรมจะจัดให้เองทุกครั้งที่เขียนทะเบียน ไม่ต้องสั่งซ้ำ

วิธีใช้:
  python printfmt.py            ทำให้ครบทุกอย่างข้างบน
  python printfmt.py pdf        เอาแค่ไฟล์ PDF พร้อมพิมพ์ ไม่แตะชีต
  python printfmt.py links      เอาแค่ลิงก์สั่งพิมพ์ของกูเกิล
  python printfmt.py --font "TH Sarabun New"    บังคับใช้ฟอนต์อื่นบนชีต
"""
import os
import sys
from urllib.parse import urlencode

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import core
import store

# ฟอนต์คนละชื่อกันสองที่ ทั้งที่เป็นแบบอักษรเดียวกัน:
#   Google Sheets  ใช้ "Sarabun" — คือ TH Sarabun ฉบับที่กูเกิลมีในระบบ
#                  ถ้าใส่ "TH Sarabun New" ไป หน้าจอเห็นถูก (เครื่องเรามีฟอนต์)
#                  แต่ PDF ที่กูเกิลสร้างจะกลายเป็น Arial เพราะฝั่งกูเกิลไม่มีฟอนต์นี้
#                  — ลองมาแล้ว ทะเบียนรับกลายเป็น ๙๑ หน้า เพราะ Arial กว้างกว่ามาก
#   ไฟล์ .xlsx     ใช้ "TH Sarabun New" ตามที่ลงไว้ในวินโดวส์ Excel จึงเห็นถูกต้อง
FONT = "Sarabun"
FONT_XLSX = "TH Sarabun New"
SIZE = 16

# ความกว้างคอลัมน์เป็นพิกเซล (๙๖ พิกเซล = ๑ นิ้ว)
# A4 แนวนอนกว้าง ๒๙๗ มม. หักขอบข้างละ ๑๐ มม. เหลือที่พิมพ์ ๒๗๗ มม. ≈ ๑,๐๔๖ พิกเซล
# รวมทุกคอลัมน์ต้องได้เท่านี้พอดี ไม่งั้นจะถูกย่อจนตัวหนังสือเล็กกว่า ๑๖
#
# แบ่งความกว้างตามข้อมูลจริงในทะเบียน (วัดความยาวข้อความทุกแถวแล้วลองจัดหน้าดู):
# ความสูงของแถวมาจากช่องที่ตัดคำแล้วยาวที่สุด ซึ่งเกือบทุกแถวคือ "เรื่อง"
# จึงยกความกว้างให้ "เรื่อง" มากที่สุด ส่วนช่องที่สั้นและซ้ำเดิมทุกแถว
# ("ถึง" = ชื่อโรงเรียน, วันที่, "การปฏิบัติ"/"หมายเหตุ" ที่ไม่เคยมีใครกรอก)
# บีบให้แคบพอให้ตัดคำได้ "๒ บรรทัด" พอดี
#
# ตัวเลข ๒ บรรทัดนี้สำคัญกว่าที่คิด: ความสูงแถวคือช่องที่บรรทัดเยอะที่สุด
# ถ้ามีช่องแคบไปแค่ช่องเดียวจนล้นเป็น ๓ บรรทัด ทุกแถวในทะเบียนจะสูงขึ้นตามทันที
# (ลองมาแล้ว — "ถึง" แคบไป ๑๐ พิกเซล ชื่อโรงเรียนล้นเป็น ๓ บรรทัด ทั้งเล่มบวมจาก
#  ๔๘ เป็น ๖๑ หน้า) ก่อนแก้ตัวเลขพวกนี้จึงควรสั่งพิมพ์ดูจริงทุกครั้ง
#
# อีกข้อ: ช่องที่ข้อมูลว่างตลอด ("การปฏิบัติ" "หมายเหตุ") ยังต้องกว้างพอให้ "หัวตาราง"
# ตัดคำได้ ไม่งั้นหัวตารางจะขาดกลางคำเป็น "การป/ฏิบัติ" ทุกหน้า
# ตัวเลขขั้นต่ำวัดจากความกว้างคำที่ยาวที่สุดของหัวตารางนั้น + ระยะขอบช่อง
#
# ผลที่วัดได้: ทะเบียนรับทั้งปีจาก ๙๑ หน้า เหลือ ๕๔ หน้า ตัวหนังสือยังขนาด ๑๖ เท่าเดิม
COLS = {
    #         เลขทะเบียน  ที่  ลงวันที่ จาก  ถึง  เรื่อง การปฏิบัติ หมายเหตุ วันที่ลงรับ
    "recv": {"widths": [75, 110, 70, 145, 125, 325, 65, 61, 70],    # รวม ๑,๐๔๖
             "center": (0, 2, 8)},          # เลขทะเบียนรับ / ลงวันที่ / วันที่ลงรับ
    #         เลขทะเบียน  ที่  ลงวันที่ จาก  ถึง  เรื่อง หมายเหตุ
    "send": {"widths": [75, 145, 140, 50, 165, 391, 80],            # รวม ๑,๐๔๖
             "center": (0, 2)},             # เลขทะเบียนส่ง / ลงวันที่
}
HEADERS_OF = {"recv": store.HEADERS, "send": store.SEND_HEADERS}
SCHOOL = store.SCHOOL
GREY = {"red": 0.925, "green": 0.925, "blue": 0.925}


def kind_of(base: str) -> str:
    """ทะเบียนเล่มไหน — ดูจากชื่อแท็บ/ชื่อไฟล์"""
    return "send" if "ส่ง" in str(base) else "recv"


# ==========================================================
# Google Sheets
# ==========================================================
def tab_requests(sheet_id: int, kind: str, font=FONT, size=SIZE) -> list:
    """คำสั่งจัดหน้าหนึ่งแท็บ (หลายแท็บส่งรวมกันทีเดียวได้ด้วย batchUpdate)"""
    spec = COLS[kind]
    ncols = len(HEADERS_OF[kind])
    whole = {"sheetId": sheet_id, "startColumnIndex": 0, "endColumnIndex": ncols}
    reqs = [
        # ตรึงหัวตาราง — บนจอเลื่อนแล้วไม่หาย บนกระดาษซ้ำทุกหน้า (ลิงก์พิมพ์ตั้ง fzr=true)
        {"updateSheetProperties": {
            "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}},
            "fields": "gridProperties.frozenRowCount"}},
        # จัดทั้งคอลัมน์ ไม่จำกัดจำนวนแถว — แถวที่เพิ่มทีหลังได้ฟอนต์นี้เอง ไม่ต้องสั่งซ้ำ
        {"repeatCell": {
            "range": whole,
            "cell": {"userEnteredFormat": {
                "textFormat": {"fontFamily": font, "fontSize": size, "bold": False},
                "wrapStrategy": "WRAP",          # เรื่องยาวขึ้นบรรทัดใหม่ ไม่ถูกตัดหาย
                "verticalAlignment": "TOP",
                "horizontalAlignment": "LEFT"}},
            "fields": "userEnteredFormat(textFormat,wrapStrategy,verticalAlignment,"
                      "horizontalAlignment)"}},
        # หัวตาราง
        {"repeatCell": {
            "range": dict(whole, startRowIndex=0, endRowIndex=1),
            "cell": {"userEnteredFormat": {
                "textFormat": {"fontFamily": font, "fontSize": size, "bold": True},
                "wrapStrategy": "WRAP",
                "verticalAlignment": "MIDDLE",
                "horizontalAlignment": "CENTER",
                "backgroundColor": GREY}},
            "fields": "userEnteredFormat(textFormat,wrapStrategy,verticalAlignment,"
                      "horizontalAlignment,backgroundColor)"}},
    ]
    # ช่องเลขและวันที่จัดกึ่งกลาง อ่านง่ายกว่าชิดซ้าย (เริ่มแถว ๒ ไม่ทับหัวตาราง)
    reqs += [{"repeatCell": {
        "range": {"sheetId": sheet_id, "startRowIndex": 1,
                  "startColumnIndex": c, "endColumnIndex": c + 1},
        "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER"}},
        "fields": "userEnteredFormat.horizontalAlignment"}} for c in spec["center"]]
    reqs += [{"updateDimensionProperties": {
        "range": {"sheetId": sheet_id, "dimension": "COLUMNS",
                  "startIndex": i, "endIndex": i + 1},
        "properties": {"pixelSize": w}, "fields": "pixelSize"}}
        for i, w in enumerate(spec["widths"])]
    # ความสูงแถวให้พอดีข้อความที่ตัดคำแล้ว (แถวที่เคยถูกตั้งความสูงตายตัวไว้ก็ปลดออก)
    reqs.append({"autoResizeDimensions": {
        "dimensions": {"sheetId": sheet_id, "dimension": "ROWS"}}})
    return reqs


def sheet_ids(st) -> dict:
    """ชื่อแท็บ -> รหัสแท็บ"""
    meta = st._run(st._api.get(spreadsheetId=st.sheet_id))
    return {s["properties"]["title"]: s["properties"]["sheetId"]
            for s in meta.get("sheets", [])}


def registry_tabs(st) -> list:
    """แท็บทะเบียนทั้งหมด ทั้งของปีนี้และปีก่อนๆ"""
    tabs = []
    for base in (st.TAB_REG, st.TAB_SEND):
        for t in (st._tabs_of(base) or [base]):
            if t not in tabs:
                tabs.append(t)
    return tabs


def format_sheets(st=None, font=FONT, size=SIZE, tabs=None) -> list:
    """จัดหน้าแท็บทะเบียนบน Google Sheets — คืนชื่อแท็บที่จัดแล้ว"""
    st = st or store.SheetsStore()
    ids = sheet_ids(st)
    todo = [t for t in (tabs or registry_tabs(st)) if t in ids]
    reqs = []
    for t in todo:
        reqs += tab_requests(ids[t], kind_of(t), font, size)
    if reqs:
        st._run(st._api.batchUpdate(spreadsheetId=st.sheet_id, body={"requests": reqs}))
    return todo


def format_new_tab(st, tab: str, base: str):
    """เรียกตอนขึ้นปีใหม่แล้วสร้างแท็บใหม่ — แท็บใหม่พร้อมพิมพ์ตั้งแต่แถวแรก"""
    sid = sheet_ids(st).get(tab)
    if sid is None:
        return
    st._run(st._api.batchUpdate(spreadsheetId=st.sheet_id, body={
        "requests": tab_requests(sid, kind_of(base))}))


# ---- ลิงก์สั่งพิมพ์ ----
# ค่าหน้ากระดาษทั้งหมดฝากไปกับลิงก์ เพราะ API ตั้งไว้ในชีตไม่ได้
PRINT_OPTS = {
    "format": "pdf",
    "size": "A4",
    "portrait": "false",        # แนวนอน
    "fitw": "true",             # บีบให้พอดีความกว้างกระดาษ กันคอลัมน์ตกขอบ
    "gridlines": "true",        # ตีเส้นตารางให้ รวมถึงแถวที่เพิ่งเพิ่มเข้ามา
    "printtitle": "true",       # ชื่อสเปรดชีตหัวกระดาษ
    "sheetnames": "true",       # ชื่อแท็บ เช่น "ทะเบียนรับ ๒๕๗๐"
    "pagenum": "CENTER",        # เลขหน้าท้ายกระดาษ
    "fzr": "true",              # หัวตารางซ้ำทุกหน้า
    "horizontal_alignment": "CENTER",
    "vertical_alignment": "TOP",
    "top_margin": "0.4", "bottom_margin": "0.4",
    "left_margin": "0.4", "right_margin": "0.4",   # นิ้ว (~๑๐ มม.)
    "attachment": "false",      # เปิดดูในเบราว์เซอร์ก่อน ไม่โหลดลงเครื่องทันที
}


def print_url(sheet_id: str, gid) -> str:
    return (f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?"
            + urlencode(dict(PRINT_OPTS, gid=gid)))


def print_links(st=None) -> list:
    """[(ชื่อแท็บ, ลิงก์พิมพ์), ...]"""
    st = st or store.SheetsStore()
    ids = sheet_ids(st)
    return [(t, print_url(st.sheet_id, ids[t])) for t in registry_tabs(st) if t in ids]


# ==========================================================
# เลขไทยทั้งทะเบียน
# ==========================================================
# ทะเบียนที่ใช้อยู่มีเลขปนกันสองแบบ ("๗๘" กับ "77", "๓๑ ส.ค. ๒๕๖๙" กับ "21 พฤษภาคม ๒๕๖๙")
# เพราะแถวเก่าพิมพ์มือ ส่วนทะเบียนส่งโปรแกรมเคยเขียนเป็นเลขอารบิก
# พิมพ์ออกมาแล้วดูไม่เรียบร้อย จึงกวาดให้เป็นเลขไทยทั้งแผ่น
#
# โปรแกรมอ่านเลขผ่าน core.to_arabic_digits() อยู่แล้วทุกที่ (หาเลขรับ ตัดปี ตรวจเลขซ้ำ)
# เปลี่ยนเป็นเลขไทยจึงไม่กระทบการทำงาน และยังทำให้การตรวจเลขซ้ำแม่นขึ้น
# เพราะเทียบสตริงตรงๆ ไม่มีกรณี "๗๗" กับ "77" ที่เป็นเลขเดียวกันแต่ไม่เท่ากัน
def _backup(data: dict) -> str:
    """เก็บของเดิมไว้ก่อนแก้ — ข้อมูลราชการ ห้ามแก้แล้วย้อนไม่ได้"""
    import json
    name = f"สำรองทะเบียน_{core.now_th().strftime('%Y%m%d_%H%M%S')}.json"
    for folder in (core.OUTPUT_ROOT, core.WORK_DIR):
        try:
            os.makedirs(folder, exist_ok=True)
            path = os.path.join(folder, name)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=1)
            return path
        except Exception:
            continue
    return ""


def thai_digits_sheets(st=None, tabs=None) -> tuple:
    """แปลงเลขอารบิกในทะเบียนบน Sheets เป็นเลขไทย — คืน ({แท็บ: จำนวนช่องที่แก้}, ไฟล์สำรอง)"""
    st = st or store.SheetsStore()
    from openpyxl.utils import get_column_letter
    before, counts = {}, {}
    for tab in tabs or registry_tabs(st):
        ncols = len(HEADERS_OF[kind_of(tab)])
        rng = f"A:{get_column_letter(ncols)}"
        vals = st._get(tab, rng, fresh=True)
        before[tab] = vals
        new = [[core.to_thai_digits(c) for c in row] for row in vals]
        n = sum(1 for a, b in zip(vals, new) for x, y in zip(a, b) if x != y)
        if not n:
            continue
        counts[tab] = n
        # เขียนกลับทั้งก้อนทีเดียว — แถวสั้นกว่าคนอื่นก็เขียนเท่าที่มี ไม่ไปลบช่องข้างหลัง
        st._run(st._api.values().update(
            spreadsheetId=st.sheet_id, range=f"'{tab}'!A1",
            valueInputOption="RAW", body={"values": new}))
    st._drop_cache()
    return counts, (_backup(before) if counts else "")


# ==========================================================
# สร้าง PDF พร้อมพิมพ์เอง — ตัดคำไทยไม่ให้คำขาด
# ==========================================================
# ทำไมต้องสร้างเอง ทั้งที่กูเกิลก็สั่งพิมพ์ได้:
#   ภาษาไทยเขียนติดกันไม่มีเว้นวรรค กูเกิลจึงตัดขึ้นบรรทัดใหม่ตรงไหนก็ได้
#   ออกมาเป็น "เลขทะเบี/ยนรับ" "กลุ่มส่งเสริมการจัดการศึ/กษา" อ่านแล้วสะดุด
#   และสั่งให้ตัดตามคำไม่ได้เลย — ไม่มีทั้งใน API และในหน้าตั้งค่าการพิมพ์
#
#   ตรงนี้จึงวาดหน้ากระดาษเอง แล้วขึ้นบรรทัดใหม่ที่ "ขอบคำ" ด้วยตัวตัดคำไทย
#   ตัวเดียวกับที่ใช้พิมพ์คำเกษียณลงหนังสือ (thaiwords + คลังคำ ๖๒,๐๙๖ คำ)
#   ได้ฟอนต์ TH Sarabun New ของจริงด้วย ไม่ใช่ฟอนต์แทนของกูเกิล
PT = 72.0                      # ๑ นิ้ว = ๗๒ จุด
PAGE_PT = (842.0, 595.0)       # A4 แนวนอน
MARGIN_PT = 28.8               # ขอบ ๐.๔ นิ้ว เท่ากับที่ตั้งไว้ในลิงก์พิมพ์


def _font(px, bold=False):
    return core.get_font(px, "THSarabunNew Bold.ttf" if bold else "THSarabunNew.ttf")


# ชื่อเฉพาะที่คลังคำทั่วไปไม่มี ถ้าไม่บอกไว้จะโดนตัดกลางชื่อ
# ("โพนทอง" คลังคำรู้จักแยกเป็น "โพน"+"ทอง" ชื่อโรงเรียนเลยขาดกลางทุกแถว)
# เรียงคำยาวไว้ก่อน จะได้จับคำยาวสุดที่ตรงก่อน
NO_BREAK = ("โพนทองประชาอุทิศ", "โพนทองวัฒนา", "โพนทอง", "สกลนคร")


def _merge_names(toks: list) -> list:
    """รวมคำที่ถูกตัดแยกให้กลับเป็นชื่อเดียว"""
    out, i = [], 0
    while i < len(toks):
        hit = None
        for kw in NO_BREAK:
            acc, j = "", i
            while j < len(toks) and len(acc) < len(kw):
                acc += toks[j]
                j += 1
            if acc == kw:
                hit = (acc, j)
                break
        if hit:
            out.append(hit[0])
            i = hit[1]
        else:
            out.append(toks[i])
            i += 1
    return out


def wrap_thai(text, font, width_px) -> list:
    """ตัดข้อความเป็นบรรทัดโดยไม่ให้คำขาดกลางคำ"""
    text = str(text or "").strip()
    if not text:
        return [""]
    lines, cur = [], ""
    for w in _merge_names(core.cached_tokenize(text)):
        if font.getlength(cur + w) <= width_px or not cur.strip():
            cur += w
            # คำเดียวยาวเกินช่อง (เช่นเลขที่หนังสือยาวๆ) ต้องหั่นทีละตัว ไม่งั้นล้นออกนอกช่อง
            while font.getlength(cur) > width_px and len(cur) > 1:
                keep = cur
                while font.getlength(keep) > width_px and len(keep) > 1:
                    keep = keep[:-1]
                lines.append(keep)
                cur = cur[len(keep):]
        else:
            lines.append(cur.rstrip())
            cur = w.lstrip()
    if cur.strip() or not lines:
        lines.append(cur.rstrip())
    return lines


def build_pdf(kind: str, rows: list, path: str, title="", dpi=200, size=SIZE) -> int:
    """วาดทะเบียนลงกระดาษ A4 แนวนอนแล้วเซฟเป็น PDF — คืนจำนวนหน้า

    วาดเป็นภาพขาวดำ (โหมด "1") แล้วให้ PDF บีบอัดแบบ CCITT G4 เหมือนโทรสาร
    ตัวหนังสือคมเท่าเดิมแต่ไฟล์เล็กกว่าภาพสีสิบเท่า
    """
    from PIL import Image, ImageDraw

    sc = dpi / PT                                   # จุด -> พิกเซล
    W, H = (int(round(x * sc)) for x in PAGE_PT)
    m = int(round(MARGIN_PT * sc))
    fs = int(round(size * sc))                      # ตัวอักษร ๑๖ จุด
    font, bold = _font(fs), _font(fs, True)
    small = _font(int(fs * 0.85))
    line_h = int(round(fs * 1.28))
    pad = max(2, int(round(2 * sc)))
    headers = HEADERS_OF[kind]
    center = set(COLS[kind]["center"])
    # ความกว้างคอลัมน์ชุดเดียวกับบนชีต (พิกเซลของชีต ๙๖ dpi -> จุด -> พิกเซลภาพ)
    cols = [int(round(w * 0.75 * sc)) for w in COLS[kind]["widths"]]
    over = sum(cols) - (W - 2 * m)
    if over > 0:                                    # ปัดเศษแล้วเกินขอบ ตัดจากช่อง "เรื่อง"
        cols[5] -= over

    def cell_lines(text, i, f):
        return wrap_thai(text, f, cols[i] - 2 * pad)

    head_lines = [cell_lines(h, i, bold) for i, h in enumerate(headers)]
    head_h = max(len(l) for l in head_lines) * line_h + 2 * pad
    body = [[cell_lines(r[i] if i < len(r) else "", i, font) for i in range(len(headers))]
            for r in rows]
    heights = [max(len(c) for c in row) * line_h + 2 * pad for row in body]

    top = m + line_h + pad                          # เว้นที่ให้หัวกระดาษหนึ่งบรรทัด
    bottom = H - m
    # แบ่งหน้า: ใส่แถวลงไปจนกว่าจะไม่พอ
    pages, cur, y = [], [], top + head_h
    for i, h in enumerate(heights):
        if cur and y + h > bottom:
            pages.append(cur); cur, y = [], top + head_h
        cur.append(i); y += h
    if cur:
        pages.append(cur)

    def draw_row(d, y, cells, f, fill_head=False):
        h = max(len(c) for c in cells) * line_h + 2 * pad
        if fill_head:
            d.rectangle([m, y, m + sum(cols), y + h], fill=230)
        x = m
        for i, lines in enumerate(cells):
            for k, ln in enumerate(lines):
                tx = x + pad
                if i in center or fill_head:
                    tx = x + (cols[i] - f.getlength(ln)) / 2
                d.text((tx, y + pad + k * line_h), ln, font=f, fill=0)
            x += cols[i]
        # เส้นตาราง
        d.rectangle([m, y, m + sum(cols), y + h], outline=0, width=max(1, int(sc / 2)))
        x = m
        for w in cols[:-1]:
            x += w
            d.line([x, y, x, y + h], fill=0, width=max(1, int(sc / 2)))
        return h

    imgs = []
    for n, page in enumerate(pages, start=1):
        im = Image.new("L", (W, H), 255)
        d = ImageDraw.Draw(im)
        d.text((m, m), title, font=small, fill=0)
        foot = f"หน้า {core.to_thai_digits(n)}/{core.to_thai_digits(len(pages))}"
        d.text((m + sum(cols) - small.getlength(foot), m), foot, font=small, fill=0)
        y = top
        y += draw_row(d, y, head_lines, bold, fill_head=True)   # หัวตารางซ้ำทุกหน้า
        for i in page:
            y += draw_row(d, y, body[i], font)
        imgs.append(im.convert("1"))                # ขาวดำ -> บีบอัดแบบโทรสาร ไฟล์เล็ก

    imgs[0].save(path, save_all=True, append_images=imgs[1:], resolution=dpi)
    return len(imgs)


def registry_data(st=None) -> list:
    """ข้อมูลทะเบียนทั้งหมด [(ชื่อ, kind, แถว), ...] — อ่านได้ทั้งบน Sheets และไฟล์ในเครื่อง"""
    from openpyxl.utils import get_column_letter
    out = []
    if (st or store.get_store()).kind == "sheets":
        st = st if isinstance(st, store.SheetsStore) else store.SheetsStore()
        for tab in registry_tabs(st):
            kind = kind_of(tab)
            rng = f"A:{get_column_letter(len(HEADERS_OF[kind]))}"
            rows = [r for r in st._get(tab, rng)[1:] if r and str(r[0]).strip()]
            out.append((tab, kind, rows))
        return out
    from openpyxl import load_workbook
    for path in (core.REGISTRY_XLSX,
                 os.path.join(core.OUTPUT_ROOT, store.SEND_XLSX_NAME)):
        if not os.path.exists(path):
            continue
        kind = kind_of(os.path.basename(path))
        n = len(HEADERS_OF[kind])
        for ws in load_workbook(path, read_only=True).worksheets:
            rows = [[c if c is not None else "" for c in r]
                    for r in ws.iter_rows(min_row=2, max_col=n, values_only=True)]
            out.append((ws.title, kind, [r for r in rows if str(r[0]).strip()]))
    return out


def make_pdfs(st=None, folder=None, dpi=200, size=SIZE) -> list:
    """สร้าง PDF พร้อมพิมพ์ของทุกทะเบียน — คืน [(ไฟล์, จำนวนหน้า, จำนวนเรื่อง)]"""
    folder = folder or core.OUTPUT_ROOT
    os.makedirs(folder, exist_ok=True)
    done = []
    for name, kind, rows in registry_data(st):
        if not rows:
            continue
        title = f"{'ทะเบียนหนังสือส่ง' if kind == 'send' else 'ทะเบียนหนังสือรับ'}" \
                f" {name.split(' ', 1)[1] if ' ' in name else ''} {SCHOOL}".replace("  ", " ")
        path = os.path.join(folder, f"{name} พร้อมพิมพ์.pdf")
        pages = build_pdf(kind, rows, path, title=title.strip(), dpi=dpi, size=size)
        done.append((path, pages, len(rows)))
    return done


# ==========================================================
# ไฟล์ .xlsx ในเครื่อง (โหมด local)
# ==========================================================
def style_ws(ws, kind: str, font=FONT_XLSX, size=SIZE):
    """จัดหน้าชีตหนึ่งชีตในไฟล์ .xlsx — ไฟล์เก็บค่าหน้ากระดาษไว้ในตัวเองได้"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties

    spec = COLS[kind]
    ncols = len(HEADERS_OF[kind])

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0          # ยาวกี่หน้าก็ได้ ขอแค่กว้างพอดีหน้าเดียว
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                                  header=0.3, footer=0.3)
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:1"            # หัวตารางซ้ำทุกหน้า
    ws.oddFooter.center.text = "หน้า &P/&N"
    ws.freeze_panes = "A2"

    for i, w in enumerate(spec["widths"], start=1):
        ws.column_dimensions[get_column_letter(i)].width = round(w / 7.0, 2)

    head_fill = PatternFill("solid", fgColor="ECECEC")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ncols):
        head = row[0].row == 1
        for c, cell in enumerate(row):
            cell.font = Font(name=font, size=size, bold=head)
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center" if head else "top",
                horizontal="center" if (head or c in spec["center"]) else "left")
            if head:
                cell.fill = head_fill


def style_workbook(wb, kind: str, font=FONT_XLSX, size=SIZE):
    for ws in wb.worksheets:
        style_ws(ws, kind, font, size)


def thai_digits_workbook(wb, kind: str) -> int:
    """เปลี่ยนเลขในไฟล์ .xlsx เป็นเลขไทย — ไม่แตะสูตร (คอลัมน์ "ที่" ในไฟล์เดิมเป็นสูตร)"""
    ncols = len(HEADERS_OF[kind])
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, max_col=ncols):
            for cell in row:
                v = cell.value
                if v is None or (isinstance(v, str) and v.startswith("=")):
                    continue
                s = core.to_thai_digits(v)
                if s != v:
                    cell.value, n = s, n + 1
    return n


def format_local(font=FONT_XLSX, size=SIZE) -> list:
    """จัดหน้า+แปลงเลขไทยให้ไฟล์ทะเบียน .xlsx ในเครื่อง — คืน [(ไฟล์, จำนวนช่องที่แปลง)]"""
    from openpyxl import load_workbook
    done = []
    for path in (core.REGISTRY_XLSX,
                 os.path.join(core.OUTPUT_ROOT, store.SEND_XLSX_NAME)):
        if not os.path.exists(path):
            continue
        kind = kind_of(os.path.basename(path))
        wb = load_workbook(path)
        style_workbook(wb, kind, font, size)
        n = thai_digits_workbook(wb, kind)
        store.LocalStore()._save_ws(wb, path)     # เซฟแบบปลอดภัย มีสำรองให้
        done.append((path, n))
    return done


# ==========================================================
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    args = sys.argv[1:]
    font, font_xlsx, size = FONT, FONT_XLSX, SIZE
    if "--font" in args:
        i = args.index("--font")
        font = font_xlsx = args[i + 1]        # สั่งมาเองก็ใช้ตัวเดียวกันทั้งสองที่
        del args[i:i + 2]
    cmd = args[0] if args else ""
    only_links = cmd == "links"
    only_pdf = cmd == "pdf"

    if not only_links:
        print("กำลังวาดไฟล์ PDF พร้อมพิมพ์ (ตัดคำไทยไม่ให้คำขาด) — ใช้เวลาสักครู่")
        for path, pages, n in make_pdfs():
            print(f"  {path}")
            print(f"    {core.to_thai_digits(n)} เรื่อง | {core.to_thai_digits(pages)} หน้า"
                  f" | A4 แนวนอน | TH Sarabun New ขนาด {core.to_thai_digits(size)}")
        print("  เปิดไฟล์แล้วสั่งพิมพ์ได้เลย ไม่ต้องตั้งค่าหน้ากระดาษอีก")
        if only_pdf:
            return

    cfg = core.load_config()
    if os.environ.get("SARABAN_SHEET_ID") or cfg.get("sheet_id"):
        st = store.SheetsStore()
        if not only_links:
            print("")
            tabs = format_sheets(st, font, size)
            print(f"จัดหน้าบน Google Sheets แล้ว {len(tabs)} แท็บ: {', '.join(tabs)}")
            print(f"  ตัวอักษร {font} ขนาด {size} | หัวตารางตัวหนา ตรึงไว้แถวบน")
            counts, bak = thai_digits_sheets(st, tabs)
            total = sum(counts.values())
            print(f"  แปลงเป็นเลขไทยแล้ว {core.to_thai_digits(total)} ช่อง"
                  + (f" ({', '.join(f'{t} {core.to_thai_digits(n)}' for t, n in counts.items())})"
                     if counts else " — เป็นเลขไทยอยู่แล้วทั้งหมด"))
            if bak:
                print(f"  ของเดิมสำรองไว้ที่ {bak}")
        print("\nลิงก์สั่งพิมพ์จากกูเกิลโดยตรง (ไว้เปิดดูจากมือถือ ไม่ต้องรอวาดไฟล์)")
        print("ข้อเสียคือกูเกิลตัดขึ้นบรรทัดใหม่กลางคำ ถ้าจะพิมพ์จริงให้ใช้ไฟล์ PDF ข้างบน")
        for tab, url in print_links(st):
            print(f"\n  {tab}\n  {url}")
    else:
        print("ยังไม่ได้ตั้งรหัสสเปรดชีต — ข้ามส่วน Google Sheets")

    if not only_links:
        done = format_local(font_xlsx, size)
        if done:
            print("\nจัดหน้าไฟล์ในเครื่องแล้ว (ค่าหน้ากระดาษ A4 แนวนอน ฝังไว้ในไฟล์):")
            for path, n in done:
                print(f"  {path} — แปลงเลขไทย {core.to_thai_digits(n)} ช่อง")


if __name__ == "__main__":
    main()

# หมายเหตุเรื่องฟอนต์ (ทำไมชีตถึงใช้ชื่อ "Sarabun" ไม่ใช่ "TH Sarabun New")
# ----------------------------------------------------------------------
# ตอนแรกใส่ "TH Sarabun New" ลงชีตตรงๆ บนจอเห็นถูกต้อง (เพราะเครื่องเรามีฟอนต์)
# แต่พอสั่งพิมพ์ PDF กลับได้ Arial เพราะเซิร์ฟเวอร์กูเกิลไม่มีฟอนต์นี้
# Arial กว้างกว่ามาก ทะเบียนบวมจาก ๔๘ หน้าเป็น ๙๑ หน้า
#
# "Sarabun" คือแบบอักษรเดียวกันฉบับที่กูเกิลมีในระบบ ใส่ชื่อนี้แล้ว PDF ฝังฟอนต์จริง
# ตัวหนังสือหน้าตาเหมือน TH Sarabun New ที่ใช้ในหนังสือราชการทุกประการ
#
# ไฟล์ .xlsx ยังใช้ "TH Sarabun New" ตามเดิม เพราะเปิดจาก Excel ในเครื่องที่มีฟอนต์อยู่แล้ว
# อยากบังคับใช้ชื่อเดียวกันทั้งสองที่:  python printfmt.py --font "TH Sarabun New"

