"""store.py — ที่เก็บข้อมูลกลาง ให้ทุกเครื่องอ่าน/เขียนที่เดียวกัน

ปัญหาที่แก้
-----------
เดิมข้อมูลอยู่เป็นไฟล์ข้างโปรแกรม (history_ids.txt, ทะเบียน .xlsx) ซึ่งพังสองแบบ:

  ๑. หลายเครื่อง — คอมกับเว็บต่างคนต่างมีสำเนา จะดึงหนังสือซ้ำ
     และที่หนักกว่าคือ "เลขรับซ้ำกัน" เพราะต่างคนต่างนับจากสำเนาตัวเอง
  ๒. hosting ฟรี — ดิสก์ถูกล้างทุกครั้งที่เซิร์ฟเวอร์รีสตาร์ท เลขรับจะเด้งกลับเป็น ๑

จึงรวบทุกอย่างมาไว้หลังหน้ากากเดียว แล้วเลือกที่เก็บจริงได้ ๒ แบบ:

  local   ไฟล์ในเครื่อง (เหมือนเดิม) แต่ล็อกไฟล์กันเขียนชนกัน
          -> เหมาะเมื่อรันเซิร์ฟเวอร์บนคอมเครื่องเดียวกับที่ใช้โปรแกรมเดสก์ท็อป
  sheets  Google Sheets
          -> เหมาะเมื่อรันบน hosting ฟรี หรือใช้หลายเครื่อง/หลายคน

เลือกด้วยตัวแปรระบบ  SARABAN_STORE=local | sheets   (ไม่ตั้ง = local)
"""
import os
import re
import threading
from datetime import datetime

import core

# ล็อกในโปรเซสเดียวกัน (เว็บมีหลาย request พร้อมกัน)
_mem_lock = threading.RLock()

HEADERS = ['เลขทะเบียนรับ', 'ที่', 'ลงวันที่', 'จาก', 'ถึง', 'เรื่อง',
           'การปฏิบัติ', 'หมายเหตุ', 'วันที่ลงรับ']
SCHOOL = "โรงเรียนบ้านโพนทองประชาอุทิศ"


def _next_no(last_value) -> str:
    """เลขรับถัดไปจากค่าล่าสุด (รองรับทั้งเลขไทยและอารบิก)"""
    try:
        return core.to_thai_digits(int(core.to_arabic_digits(str(last_value))) + 1)
    except (ValueError, TypeError):
        return core.to_thai_digits(1)


def _row(receipt_no, doc_no, doc_date, sender, doc_title, receive_date, thai_date_fn):
    """สร้างแถวทะเบียนตามลำดับคอลัมน์มาตรฐาน"""
    return [receipt_no,
            core.to_thai_digits((doc_no or "").strip()),
            core.normalize_typed_date(doc_date or ""),
            (sender or "").strip(),
            SCHOOL,
            (doc_title or "").strip(),
            "", "",
            core.normalize_typed_date(receive_date or "") or thai_date_fn()]


# คอลัมน์ของแท็บผู้ใช้ (เก็บเป็นคอลัมน์ ไม่ใช่ JSON ก้อนเดียว จะได้เปิดดูในชีตรู้เรื่อง)
USER_COLS = ['username', 'salt', 'hash', 'display', 'role', 'status', 'created', 'must_change']


class UsersMixin:
    """เก็บรายชื่อผู้ใช้เว็บ

    บน hosting ฟรี ดิสก์ถูกล้างทุกครั้งที่รีสตาร์ท ถ้าเก็บ users.json ไว้บนดิสก์
    ผู้ใช้จะหายหมดและต้องสมัครใหม่ทุกครั้ง จึงต้องเก็บที่เดียวกับทะเบียน

    หมายเหตุ: เก็บแค่ salt กับ hash ไม่เคยเก็บรหัสผ่านจริง
    """

    @staticmethod
    def _users_from_rows(rows) -> dict:
        out = {}
        for r in rows:
            r = list(r) + [""] * (len(USER_COLS) - len(r))
            name = str(r[0]).strip()
            if not name:
                continue
            u = {k: r[i] for i, k in enumerate(USER_COLS[1:], start=1)}
            u["must_change"] = str(u.get("must_change", "")).lower() in ("true", "1", "yes")
            out[name] = u
        return out

    @staticmethod
    def _rows_from_users(users: dict) -> list:
        return [[name] + [("TRUE" if u.get("must_change") else "")
                          if k == "must_change" else str(u.get(k, ""))
                          for k in USER_COLS[1:]]
                for name, u in users.items()]


DONE = "registered"      # ลงรับแล้ว
SKIP = "skipped"         # ข้าม (ไม่รับ)
NEW = "new"              # ยังไม่ลงรับ


class StatusMixin:
    """แยกสถานะหนังสือแต่ละเรื่องว่า รับแล้ว / ข้าม / ยังไม่ลงรับ

    ของเดิมเก็บแค่รายการ ID ที่ "จัดการไปแล้ว" ไม่ได้บอกว่ารับหรือข้าม
    จึงหาสถานะย้อนหลังด้วยการจับคู่ "เลขหนังสือ" กับทะเบียนรับ
    (ทะเบียนเก็บเป็นเลขไทย เว็บให้เลขอารบิก จึงแปลงให้ตรงกันก่อนเทียบ)

    ส่วนที่ลงรับตั้งแต่นี้ไปจะบันทึกสถานะไว้ตรงๆ ไม่ต้องเดา
    """

    def _reg_index(self) -> dict:
        """แผนที่ เลขหนังสือ(อารบิก) -> เลขรับ"""
        out = {}
        for r in self.registry_rows():
            if r and len(r) > 1 and r[1]:
                out[core.to_arabic_digits(str(r[1])).strip()] = r[0]
        return out

    def status_of(self, docs: list) -> list:
        """เติมสถานะให้รายการหนังสือที่ดึงมาจากเว็บ

        ของเก่าที่อยู่ในไฟล์ประวัติอยู่แล้ว ถือว่า "รับแล้ว" ทั้งหมด ไม่ต้องเดาย้อนหลัง
        (ยังพยายามหาเลขรับจากทะเบียนมาแสดงให้ ถ้าจับคู่เลขหนังสือได้)
        เฉพาะที่ทำตั้งแต่มีระบบนี้เป็นต้นไป จึงจะแยก รับแล้ว/ข้าม ได้ชัดเจน
        """
        recs = self.doc_records()
        reg = self._reg_index()
        hist = self.history_ids()
        out = []
        for d in docs:
            d = dict(d)
            bid = str(d.get("book_id", ""))
            rec = recs.get(bid)
            if rec and rec.get("status"):
                d["status"] = rec["status"]
                d["receipt_no"] = rec.get("receipt_no", "")
            elif bid in hist:
                d["status"] = DONE
                d["receipt_no"] = reg.get(core.to_arabic_digits(d.get("doc_no", "")).strip(), "")
            else:
                d["status"], d["receipt_no"] = NEW, ""
            out.append(d)
        return out

    def mark_registered(self, book_id, receipt_no):
        self.set_record(book_id, {"status": DONE, "receipt_no": str(receipt_no),
                                  "when": datetime.now().strftime("%Y-%m-%d %H:%M")})
        self.add_history(book_id)

    def mark_skipped(self, book_id):
        self.set_record(book_id, {"status": SKIP, "receipt_no": "",
                                  "when": datetime.now().strftime("%Y-%m-%d %H:%M")})
        self.add_history(book_id)


# ==========================================================
# แบบที่ ๑ — ไฟล์ในเครื่อง (ล็อกไฟล์กันเขียนชนกัน)
# ==========================================================
class LocalStore(StatusMixin, UsersMixin):
    """เก็บเป็นไฟล์เหมือนเดิม แต่เพิ่มล็อกไฟล์

    ล็อกไฟล์สำคัญตรงที่โปรแกรมเดสก์ท็อปกับเซิร์ฟเวอร์เว็บเป็นคนละโปรเซส
    ถ้าไม่ล็อก สองฝั่งอาจอ่านเลขรับเดิมพร้อมกันแล้วได้เลขซ้ำ
    """
    kind = "local"

    def __init__(self):
        from filelock import FileLock
        os.makedirs(core.OUTPUT_ROOT, exist_ok=True)
        self._lock = FileLock(os.path.join(core.OUTPUT_ROOT, ".saraban.lock"), timeout=30)

        # ประวัติต้องอยู่โฟลเดอร์เดียวกับทะเบียน ไม่ใช่ข้างโค้ด
        # เพราะโฟลเดอร์นี้คือ "ที่กลาง" ที่ทุกเครื่องชี้มาร่วมกัน (และซิงก์ขึ้น Drive อยู่แล้ว)
        self._hist = os.path.join(core.OUTPUT_ROOT, "history_ids.txt")
        self._migrate_history()

    def _migrate_history(self):
        """ย้ายไฟล์ประวัติเดิมที่อยู่ข้างโค้ดมาไว้โฟลเดอร์กลาง (ทำครั้งเดียว)"""
        old = core._p("history_ids.txt")
        if os.path.exists(self._hist) or not os.path.exists(old):
            return
        try:
            import shutil
            shutil.copy2(old, self._hist)
            os.replace(old, old + ".migrated")   # เก็บของเดิมไว้เผื่อ ไม่ลบทิ้ง
        except Exception:
            self._hist = old                      # ย้ายไม่ได้ก็ใช้ที่เดิมต่อไป

    # ---- ประวัติหนังสือที่ดึงไปแล้ว ----
    def history_ids(self) -> set:
        if not os.path.exists(self._hist):
            return set()
        with open(self._hist, encoding="utf-8", errors="ignore") as f:
            return {ln.strip() for ln in f if ln.strip()}

    def add_history(self, book_id: str):
        book_id = str(book_id).strip()
        if not book_id:
            return
        with _mem_lock, self._lock:
            if book_id in self.history_ids():      # กันเขียนซ้ำ
                return
            with open(self._hist, "a", encoding="utf-8") as f:
                f.write(book_id + "\n")

    # ---- ทะเบียนรับ ----
    def _open_ws(self):
        from openpyxl import Workbook, load_workbook
        path = core.REGISTRY_XLSX
        if not os.path.exists(path):
            wb = Workbook(); ws = wb.active; ws.title = "ทะเบียนรับ"; ws.append(HEADERS)
        else:
            wb = load_workbook(path); ws = wb.active
        return wb, ws, path

    def _save_ws(self, wb, path=None):
        """เซฟทะเบียนแบบปลอดภัย — เขียนไฟล์ชั่วคราวให้เสร็จก่อน แล้วค่อยสลับเข้าที่

        ถ้าเขียนทับไฟล์จริงตรงๆ แล้วไฟดับ/โปรแกรมถูกปิดกลางคัน
        ทะเบียนทั้งไฟล์จะพังและกู้ไม่ได้ (นี่คือไฟล์สำคัญที่สุดของระบบ)
        วิธีนี้ไฟล์จริงจะสมบูรณ์เสมอ ไม่ว่าจะดับตอนไหน
        """
        path = path or core.REGISTRY_XLSX
        tmp = path + ".tmp"
        wb.save(tmp)
        # สำรองตัวก่อนหน้าไว้หนึ่งรุ่น เผื่อต้องย้อน
        if os.path.exists(path):
            bak = path + ".bak"
            try:
                if os.path.exists(bak):
                    os.remove(bak)
                os.replace(path, bak)
            except Exception:
                pass
        os.replace(tmp, path)     # สลับไฟล์แบบอะตอมมิก

    def peek_receipt_no(self) -> str:
        with _mem_lock, self._lock:
            return self._peek()

    def _peek(self) -> str:
        if not os.path.exists(core.REGISTRY_XLSX):
            return core.to_thai_digits(1)
        from openpyxl import load_workbook
        ws = load_workbook(core.REGISTRY_XLSX).active
        for r in range(ws.max_row, 1, -1):
            v = ws.cell(row=r, column=1).value
            if v is not None:
                try:
                    return _next_no(v)
                except Exception:
                    continue
        return core.to_thai_digits(1)

    def register(self, doc_no="", doc_date="", sender="", doc_title="", receive_date="") -> str:
        """จองเลขรับและเขียนแถวทะเบียนในล็อกเดียว — กันเลขซ้ำ"""
        with _mem_lock, self._lock:
            no = self._peek()
            wb, ws, path = self._open_ws()
            ws.append(_row(no, doc_no, doc_date, sender, doc_title, receive_date, core.get_thai_date))
            self._save_ws(wb, path)
            return no

    def register_with_no(self, receipt_no, doc_no="", doc_date="", sender="",
                         doc_title="", receive_date="") -> str:
        """เขียนแถวโดยระบุเลขรับมาเอง (ใช้ตอนจองเลขไว้ก่อนหน้าแล้ว)"""
        with _mem_lock, self._lock:
            wb, ws, path = self._open_ws()
            ws.append(_row(receipt_no, doc_no, doc_date, sender, doc_title,
                           receive_date, core.get_thai_date))
            self._save_ws(wb, path)
            return receipt_no

    # ---- บันทึกสถานะแต่ละเรื่อง ----
    def _rec_path(self):
        return os.path.join(core.OUTPUT_ROOT, "doc_records.json")

    def doc_records(self) -> dict:
        import json
        p = self._rec_path()
        if not os.path.exists(p):
            return {}
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def set_record(self, book_id, rec: dict):
        import json
        with _mem_lock, self._lock:
            data = self.doc_records()
            data[str(book_id)] = rec
            with open(self._rec_path(), "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=1)

    def update_registry_row(self, receipt_no, doc_no="", doc_date="", sender="",
                            doc_title="", receive_date="") -> bool:
        """เขียนทับแถวเดิมของเลขรับนี้ — ใช้ตอน 'ลงรับใหม่' จะได้ไม่กินเลขเพิ่ม"""
        want = core.to_arabic_digits(str(receipt_no).strip())
        with _mem_lock, self._lock:
            if not os.path.exists(core.REGISTRY_XLSX):
                return False
            from openpyxl import load_workbook
            wb = load_workbook(core.REGISTRY_XLSX); ws = wb.active
            for r in range(ws.max_row, 1, -1):
                v = ws.cell(row=r, column=1).value
                if v is not None and core.to_arabic_digits(str(v).strip()) == want:
                    row = _row(v, doc_no, doc_date, sender, doc_title,
                               receive_date, core.get_thai_date)
                    for c, val in enumerate(row, start=1):
                        if c in (7, 8):        # การปฏิบัติ/หมายเหตุ ไม่ทับของเดิม
                            continue
                        ws.cell(row=r, column=c).value = val
                    self._save_ws(wb)
                    return True
            return False

    def registry_rows(self, limit=None):
        if not os.path.exists(core.REGISTRY_XLSX):
            return []
        from openpyxl import load_workbook
        ws = load_workbook(core.REGISTRY_XLSX).active
        rows = [[c.value for c in ws[r]] for r in range(2, ws.max_row + 1)]
        return rows[-limit:] if limit else rows

    # ---- ผู้ใช้เว็บ ----
    def _users_path(self):
        return core._p("users.json")

    def load_users(self) -> dict:
        import json
        p = self._users_path()
        if not os.path.exists(p):
            return {}
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def save_users(self, users: dict):
        import json
        with _mem_lock, self._lock:
            p = self._users_path()
            tmp = p + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(users, f, ensure_ascii=False, indent=2)
            os.replace(tmp, p)        # สลับแบบอะตอมมิก กันไฟล์พังถ้าดับกลางคัน

    # ---- ถอยกลับเวลาลงรับผิด ----
    def remove_history(self, book_id: str) -> bool:
        """เอาเรื่องออกจากประวัติ เพื่อให้กลับมาขึ้นในรายการ 'หนังสือใหม่' อีกครั้ง"""
        book_id = str(book_id).strip()
        with _mem_lock, self._lock:
            ids = self.history_ids()
            if book_id not in ids:
                return False
            keep = [i for i in self._ordered_history() if i != book_id]
            with open(self._hist, "w", encoding="utf-8") as f:
                f.write("\n".join(keep) + ("\n" if keep else ""))
            return True

    def _ordered_history(self):
        if not os.path.exists(self._hist):
            return []
        with open(self._hist, encoding="utf-8", errors="ignore") as f:
            return [ln.strip() for ln in f if ln.strip()]

    def delete_receipt(self, receipt_no: str) -> dict:
        """ลบแถวทะเบียนตามเลขรับ คืนข้อมูลแถวที่ลบไว้เผื่อแสดงให้ดู

        ถ้าลบแถวสุดท้าย เลขนั้นจะถูกนำกลับมาใช้ใหม่ได้ทันที
        ถ้าลบแถวกลางๆ จะเว้นเลขนั้นว่างไว้ (เลขถัดไปยังนับต่อจากตัวมากสุดเหมือนเดิม)
        """
        want = core.to_arabic_digits(str(receipt_no).strip())
        with _mem_lock, self._lock:
            if not os.path.exists(core.REGISTRY_XLSX):
                return {}
            from openpyxl import load_workbook
            wb = load_workbook(core.REGISTRY_XLSX)
            ws = wb.active
            for r in range(ws.max_row, 1, -1):
                v = ws.cell(row=r, column=1).value
                if v is not None and core.to_arabic_digits(str(v).strip()) == want:
                    row = [c.value for c in ws[r]]
                    was_last = (r == ws.max_row)
                    ws.delete_rows(r)
                    self._save_ws(wb)
                    return {"row": row, "reusable": was_last}
            return {}


# ==========================================================
# แบบที่ ๒ — Google Sheets (ใช้ร่วมกันได้ทุกเครื่อง)
# ==========================================================
class SheetsStore(StatusMixin, UsersMixin):
    """เก็บบน Google Sheets — ทุกเครื่องอ่านเขียนที่เดียวกันจริงๆ

    ทำไมถึงเลือก Sheets แทนไฟล์บน Google Drive:
      ไฟล์ธรรมดาต้อง "โหลดมา -> แก้ -> อัปกลับ" ถ้าสองคนทำพร้อมกัน
      คนที่อัปทีหลังจะทับงานคนแรกหายไปเลย
      ส่วน Sheets มีคำสั่ง append ที่ต่อท้ายให้แบบไม่ชนกัน (ฝั่งกูเกิลจัดคิวให้)

    ต้องตั้งค่าตัวแปรระบบ:
      SARABAN_STORE=sheets
      SARABAN_SHEET_ID=<รหัสสเปรดชีต จาก URL>
      SARABAN_SA_JSON=<เนื้อไฟล์ service account JSON ทั้งก้อน>  หรือ
      SARABAN_SA_FILE=<ที่อยู่ไฟล์ JSON>
    """
    kind = "sheets"
    CACHE_TTL = 20          # วินาที — จำผลอ่านไว้สั้นๆ กันเรียก API ถี่
    TAB_REG = "ทะเบียนรับ"
    TAB_HIST = "ประวัติที่ดึงแล้ว"
    TAB_USERS = "ผู้ใช้เว็บ"
    USERS_TTL = 15          # วินาที — current_user เรียกทุก request จึงต้องแคช
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

    def __init__(self):
        import json
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build

        cfg = core.load_config()
        # ตั้งค่าได้สองทาง: ตัวแปรระบบ (สำหรับ hosting) หรือ config.json (สำหรับเครื่องนี้)
        self.sheet_id = (os.environ.get("SARABAN_SHEET_ID", "").strip()
                         or cfg.get("sheet_id", "").strip())
        if not self.sheet_id:
            raise RuntimeError("ยังไม่ได้ตั้งรหัสสเปรดชีต (sheet_id) — รัน: python setup_sheets.py check")

        raw = os.environ.get("SARABAN_SA_JSON", "").strip()
        if raw:
            info = json.loads(raw)
        else:
            path = (os.environ.get("SARABAN_SA_FILE", "").strip()
                    or cfg.get("sa_file", "").strip())
            # บน hosting ให้วางเนื้อไฟล์กุญแจทั้งก้อนไว้ใน SARABAN_SA_JSON แทน
            # (ไฟล์ .json อัปขึ้น GitHub ไม่ได้ เพราะมี private key)
            if not path or not os.path.exists(path):
                raise RuntimeError("ยังไม่พบไฟล์กุญแจ service account — รัน: python setup_sheets.py check")
            with open(path, encoding="utf-8") as f:
                info = json.load(f)
        self.client_email = info.get("client_email", "")

        creds = Credentials.from_service_account_info(info, scopes=self.SCOPES)
        self._api = build("sheets", "v4", credentials=creds, cache_discovery=False).spreadsheets()
        self._hist_cache = None
        self._hist_at = 0
        self._cache = {}
        self._users_cache = None
        self._ensure_tabs()

    # ---- เตรียมแท็บให้ครบ ----
    def _ensure_tabs(self):
        meta = self._run(self._api.get(spreadsheetId=self.sheet_id))
        have = {s["properties"]["title"] for s in meta.get("sheets", [])}
        want = [t for t in (self.TAB_REG, self.TAB_HIST, self.TAB_USERS) if t not in have]
        if want:
            self._run(self._api.batchUpdate(spreadsheetId=self.sheet_id, body={"requests": [
                {"addSheet": {"properties": {"title": t}}} for t in want]}))
        if self.TAB_REG in want:
            self._append(self.TAB_REG, [HEADERS])
        if self.TAB_HIST in want:
            self._append(self.TAB_HIST, [["book_id", "เวลาที่บันทึก", "สถานะ", "เลขรับ"]])
        if self.TAB_USERS in want:
            self._append(self.TAB_USERS, [USER_COLS])

    RETRY = 3               # ลองซ้ำกี่ครั้งเมื่อเน็ตสะดุด

    @staticmethod
    def _run(req):
        """ยิงคำสั่งไป Google พร้อมลองซ้ำเมื่อเจอปัญหาชั่วคราว

        เน็ตสะดุดแค่วินาทีเดียวก็ทำให้ทั้งระบบล่มได้ (ต่างจากตอนเก็บเป็นไฟล์
        ที่ไม่ต้องพึ่งเน็ตเลย) จึงต้องลองซ้ำก่อนยอมแพ้
        """
        import time as _t
        last = None
        for i in range(SheetsStore.RETRY):
            try:
                return req.execute()
            except Exception as e:
                last = e
                msg = f"{type(e).__name__}: {e}".lower()
                temporary = any(k in msg for k in (
                    "transporterror", "servernotfound", "unable to find the server",
                    "timed out", "timeout", "connection", "ssl",
                    "500", "502", "503", "504", "rate", "quota"))
                if not temporary or i == SheetsStore.RETRY - 1:
                    raise
                _t.sleep(1.5 * (i + 1))     # รอนานขึ้นทีละรอบ
        raise last

    def _get(self, tab, rng="A:I", fresh=False):
        """อ่านข้อมูลจากชีต — มีแคชสั้นๆ กันช้า

        ทุกครั้งที่เรียก Google API ใช้เวลา ๑–๓ วินาที ถ้าหน้าเว็บหน้าเดียว
        เรียกหลายรอบจะรอนานมาก จึงจำผลไว้แป๊บหนึ่ง
        แต่ตอน "จองเลขรับ" ต้องอ่านสดเสมอ (fresh=True) ไม่งั้นอาจได้เลขซ้ำ
        """
        import time
        key = (tab, rng)
        if not fresh:
            hit = self._cache.get(key)
            if hit and time.time() - hit[0] < self.CACHE_TTL:
                return hit[1]
        r = self._run(self._api.values().get(spreadsheetId=self.sheet_id,
                                   range=f"'{tab}'!{rng}"))
        vals = r.get("values", [])
        self._cache[key] = (time.time(), vals)
        return vals

    def _drop_cache(self):
        """ล้างแคชหลังเขียน เพื่อให้รอบถัดไปเห็นของจริง"""
        self._cache.clear()
        self._hist_cache = None
        self._users_cache = None

    def _append(self, tab, rows):
        """ต่อท้ายแบบไม่ชนกัน — INSERT_ROWS ให้กูเกิลจัดคิวให้เอง"""
        res = self._run(self._api.values().append(
            spreadsheetId=self.sheet_id, range=f"'{tab}'!A1",
            valueInputOption="RAW", insertDataOption="INSERT_ROWS",
            body={"values": rows}))
        self._drop_cache()
        return res

    # ---- ประวัติ ----
    def history_ids(self) -> set:
        import time
        now = time.time()
        if self._hist_cache is not None and now - self._hist_at < 20:
            return self._hist_cache          # กันยิง API ถี่เกินไป
        vals = self._get(self.TAB_HIST, "A:A")
        ids = {str(r[0]).strip() for r in vals[1:] if r and str(r[0]).strip()}
        self._hist_cache, self._hist_at = ids, now
        return ids

    def add_history(self, book_id: str):
        book_id = str(book_id).strip()
        if not book_id:
            return
        with _mem_lock:
            self._hist_cache = None           # ล้าง cache ให้รอบหน้าอ่านของจริง
            if book_id in self.history_ids():
                return
            self._append(self.TAB_HIST,
                         [[book_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S")]])
            self._hist_cache = None

    # ---- ทะเบียนรับ ----
    def _last_no(self, vals):
        for row in reversed(vals[1:]):
            if row and str(row[0]).strip():
                return row[0]
        return None

    def peek_receipt_no(self) -> str:
        return _next_no(self._last_no(self._get(self.TAB_REG)))

    def register(self, doc_no="", doc_date="", sender="", doc_title="", receive_date="") -> str:
        """จองเลขรับบน Sheets

        กันเลขซ้ำสองชั้น:
          ๑. ต่อท้ายด้วย INSERT_ROWS ซึ่งกูเกิลรับประกันว่าไม่ทับกัน
          ๒. อ่านกลับมาดู ถ้าเลขที่ได้ดันซ้ำกับใคร ให้เลื่อนเป็นเลขถัดไปแล้วแก้เฉพาะช่องตัวเอง
        """
        with _mem_lock:
            vals = self._get(self.TAB_REG, fresh=True)   # ต้องสด กันเลขซ้ำ
            no = _next_no(self._last_no(vals))
            row = _row(no, doc_no, doc_date, sender, doc_title, receive_date, core.get_thai_date)
            res = self._append(self.TAB_REG, [row])

            # หาว่าแถวเราไปลงที่บรรทัดไหน เช่น 'ทะเบียนรับ'!A387:I387
            m = re.search(r"!\D+(\d+)", res.get("updates", {}).get("updatedRange", ""))
            if not m:
                return no
            my_row = int(m.group(1))

            after = self._get(self.TAB_REG, "A:A", fresh=True)
            used = [str(r[0]).strip() for i, r in enumerate(after[1:], start=2)
                    if r and str(r[0]).strip() and i != my_row]
            if no in used:                    # มีคนอื่นแทรกเลขเดียวกันมาก่อน
                no = _next_no(self._last_no([[""]] + [[u] for u in used]))
                self._run(self._api.values().update(
                    spreadsheetId=self.sheet_id,
                    range=f"'{self.TAB_REG}'!A{my_row}",
                    valueInputOption="RAW", body={"values": [[no]]}))
                self._drop_cache()
            return no

    def register_with_no(self, receipt_no, doc_no="", doc_date="", sender="",
                         doc_title="", receive_date="") -> str:
        with _mem_lock:
            self._append(self.TAB_REG, [_row(receipt_no, doc_no, doc_date, sender,
                                             doc_title, receive_date, core.get_thai_date)])
            return receipt_no

    # ---- บันทึกสถานะแต่ละเรื่อง (เก็บในแท็บประวัติ คอลัมน์ C, D) ----
    def doc_records(self) -> dict:
        out = {}
        for r in self._get(self.TAB_HIST, "A:D")[1:]:
            if not r or not str(r[0]).strip():
                continue
            r = list(r) + [""] * (4 - len(r))
            if r[2]:
                out[str(r[0]).strip()] = {"status": r[2], "receipt_no": r[3], "when": r[1]}
        return out

    def set_record(self, book_id, rec: dict):
        book_id = str(book_id).strip()
        with _mem_lock:
            vals = self._get(self.TAB_HIST, "A:A")
            row_no = next((i for i, r in enumerate(vals[1:], start=2)
                           if r and str(r[0]).strip() == book_id), None)
            line = [book_id, rec.get("when", ""), rec.get("status", ""), rec.get("receipt_no", "")]
            if row_no:
                self._run(self._api.values().update(
                    spreadsheetId=self.sheet_id, range=f"'{self.TAB_HIST}'!A{row_no}:D{row_no}",
                    valueInputOption="RAW", body={"values": [line]}))
                self._drop_cache()
            else:
                self._append(self.TAB_HIST, [line])
            self._hist_cache = None

    def update_registry_row(self, receipt_no, doc_no="", doc_date="", sender="",
                            doc_title="", receive_date="") -> bool:
        want = core.to_arabic_digits(str(receipt_no).strip())
        with _mem_lock:
            vals = self._get(self.TAB_REG)
            for i in range(len(vals) - 1, 0, -1):
                r = vals[i]
                if r and r[0] and core.to_arabic_digits(str(r[0]).strip()) == want:
                    old = list(r) + [""] * (9 - len(r))
                    row = _row(r[0], doc_no, doc_date, sender, doc_title,
                               receive_date, core.get_thai_date)
                    row[6], row[7] = old[6], old[7]     # คงการปฏิบัติ/หมายเหตุเดิม
                    self._run(self._api.values().update(
                        spreadsheetId=self.sheet_id,
                        range=f"'{self.TAB_REG}'!A{i+1}:I{i+1}",
                        valueInputOption="RAW", body={"values": [row]}))
                    self._drop_cache()
                    return True
            return False

    def registry_rows(self, limit=None):
        rows = self._get(self.TAB_REG)[1:]
        return rows[-limit:] if limit else rows

    # ---- ผู้ใช้เว็บ ----
    def load_users(self) -> dict:
        import time
        now = time.time()
        c = getattr(self, "_users_cache", None)
        if c and now - c[0] < self.USERS_TTL:
            return c[1]
        users = self._users_from_rows(self._get(self.TAB_USERS, "A:H")[1:])
        self._users_cache = (now, users)
        return users

    def save_users(self, users: dict):
        """เขียนทับทั้งแท็บ (ผู้ใช้มีไม่กี่คน เขียนทีเดียวจบง่ายกว่าไล่แก้ทีละแถว)"""
        with _mem_lock:
            self._run(self._api.values().clear(
                spreadsheetId=self.sheet_id, range=f"'{self.TAB_USERS}'!A2:H10000", body={}))
            rows = self._rows_from_users(users)
            if rows:
                self._run(self._api.values().update(
                    spreadsheetId=self.sheet_id,
                    range=f"'{self.TAB_USERS}'!A2",
                    valueInputOption="RAW", body={"values": rows}))
            self._users_cache = None
            self._drop_cache()

    # ---- ถอยกลับเวลาลงรับผิด ----
    def _tab_id(self, title):
        meta = self._run(self._api.get(spreadsheetId=self.sheet_id))
        for s in meta.get("sheets", []):
            if s["properties"]["title"] == title:
                return s["properties"]["sheetId"]
        raise RuntimeError(f"ไม่พบแท็บ {title}")

    def _delete_row(self, tab, row_no):
        """ลบทั้งแถว (row_no นับแบบที่เห็นในชีต บรรทัดแรกคือ ๑)"""
        self._run(self._api.batchUpdate(spreadsheetId=self.sheet_id, body={"requests": [{
            "deleteDimension": {"range": {
                "sheetId": self._tab_id(tab), "dimension": "ROWS",
                "startIndex": row_no - 1, "endIndex": row_no}}}]}))
        self._drop_cache()

    def remove_history(self, book_id: str) -> bool:
        book_id = str(book_id).strip()
        with _mem_lock:
            vals = self._get(self.TAB_HIST, "A:A")
            for i, r in enumerate(vals[1:], start=2):
                if r and str(r[0]).strip() == book_id:
                    self._delete_row(self.TAB_HIST, i)
                    self._hist_cache = None
                    return True
            return False

    def delete_receipt(self, receipt_no: str) -> dict:
        want = core.to_arabic_digits(str(receipt_no).strip())
        with _mem_lock:
            vals = self._get(self.TAB_REG)
            for i in range(len(vals) - 1, 0, -1):
                r = vals[i]
                if r and r[0] and core.to_arabic_digits(str(r[0]).strip()) == want:
                    was_last = (i == len(vals) - 1)
                    self._delete_row(self.TAB_REG, i + 1)   # +1 เพราะ vals เริ่มนับที่ ๐
                    return {"row": r, "reusable": was_last}
            return {}


# ==========================================================
# ตัวเลือกที่เก็บ
# ==========================================================
_store = None


def get_store():
    global _store
    if _store is None:
        with _mem_lock:
            if _store is None:
                kind = (os.environ.get("SARABAN_STORE", "").strip().lower()
                        or core.load_config().get("store_mode", "").strip().lower()
                        or "local")
                _store = SheetsStore() if kind == "sheets" else LocalStore()
    return _store


def reset_store():
    """ใช้ตอนทดสอบ/สลับที่เก็บ"""
    global _store
    with _mem_lock:
        _store = None
